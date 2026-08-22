"""Evidence-document parsers — pure functions, bytes in, `ExtractedChunk`s out.

These parse a document that is being used as an *evidence source* (Main
Spec §17 Phase 2 item 3/4), as opposed to `parse.py`'s `parse_document()`,
which parses a *questionnaire*. Same purity boundary as the rest of this
package (AGENTS.md §3.2/3.3, CTO-RULINGS BLOCKER-04):

- No DB session, no HTTP client, no credentials, no filesystem access beyond
  the in-memory bytes handed in.
- Never assigns a `document_chunks.id`, never resolves or invents a
  persisted source location — only the fields needed for the SERVER to build
  one (`page_number`, `sheet_name`, `cell_range`, `heading_path`).
- Uploaded document content is untrusted data (trust boundary TB-3): it is
  only ever extracted as text here, never interpreted as an instruction.

Each function raises `ValueError` on a file it cannot parse or that yields
no chunkable content — a legitimate, catchable failure, not a crash. The
caller (apps/api) is responsible for turning that into a FAILED
`processing_jobs` row; this package never touches a database.
"""

from __future__ import annotations

import io

from .models import ExtractedChunk

# --------------------------------------------------------------------------- #
# PDF (PyMuPDF / fitz) — one chunk per page.
# --------------------------------------------------------------------------- #


def parse_pdf_evidence(file_bytes: bytes) -> list[ExtractedChunk]:
    """Extract one `ExtractedChunk` per PDF page, `page_number` set (1-based).

    Blank pages (no extractable text) are skipped. Raises `ValueError` if the
    file cannot be opened as a PDF, or if no page yields any text (e.g. a
    scanned/image-only PDF with no OCR layer — OCR fallback is out of scope
    for this pass; such a file is a legitimate parse failure here).
    """
    import fitz  # PyMuPDF

    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
    except Exception as exc:  # PyMuPDF raises its own exception types
        raise ValueError(f"could not open file as a PDF: {exc}") from exc

    chunks: list[ExtractedChunk] = []
    try:
        for page_index in range(doc.page_count):
            page = doc.load_page(page_index)
            text = page.get_text("text").strip()
            if not text:
                continue
            chunks.append(
                ExtractedChunk(
                    sequence_no=len(chunks),
                    text=text,
                    page_number=page_index + 1,
                )
            )
    finally:
        doc.close()

    if not chunks:
        raise ValueError("no extractable text found in any PDF page")

    return chunks


# --------------------------------------------------------------------------- #
# DOCX (python-docx) — one chunk per heading section. No fixed pages, so
# `page_number` stays null; `heading_path` records the section instead.
# --------------------------------------------------------------------------- #


def _heading_level(style_name: str | None) -> int | None:
    """`"Heading 1"` -> 1, `"Heading 2"` -> 2, ... `"Title"` -> 0. None if the
    paragraph style is not a heading."""
    if not style_name:
        return None
    name = style_name.strip().lower()
    if name == "title":
        return 0
    if name.startswith("heading"):
        suffix = name.replace("heading", "").strip()
        if suffix.isdigit():
            return int(suffix)
    return None


def parse_docx_evidence(file_bytes: bytes) -> list[ExtractedChunk]:
    """Extract one `ExtractedChunk` per heading section: all body paragraph
    text between one heading and the next, with `heading_path` set to the
    stack of enclosing heading titles. `page_number` is always null — DOCX
    has no fixed pages (Main Spec §17 Phase 2 item 4).

    Text before the first heading (if any) is emitted as a chunk with an
    empty `heading_path`. Raises `ValueError` if the file cannot be opened as
    a DOCX, or if the document contains no extractable paragraph text at all.
    """
    import docx  # python-docx

    try:
        document = docx.Document(io.BytesIO(file_bytes))
    except Exception as exc:
        raise ValueError(f"could not open file as a DOCX: {exc}") from exc

    chunks: list[ExtractedChunk] = []
    heading_stack: list[tuple[int, str]] = []  # (level, title)
    current_lines: list[str] = []

    def _flush() -> None:
        text = "\n".join(line for line in current_lines if line.strip()).strip()
        if text:
            chunks.append(
                ExtractedChunk(
                    sequence_no=len(chunks),
                    text=text,
                    page_number=None,
                    heading_path=[title for _, title in heading_stack],
                )
            )
        current_lines.clear()

    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        level = _heading_level(paragraph.style.name if paragraph.style else None)

        if level is not None and text:
            _flush()
            while heading_stack and heading_stack[-1][0] >= level:
                heading_stack.pop()
            heading_stack.append((level, text))
            continue

        if text:
            current_lines.append(text)

    _flush()

    if not chunks:
        raise ValueError("no extractable paragraph text found in the DOCX file")

    return chunks


# --------------------------------------------------------------------------- #
# XLSX-as-evidence (openpyxl) — one chunk per non-blank row, `sheet_name` +
# `cell_range` set. Distinct from parse.py's parse_document(), which parses
# an XLSX as a *questionnaire*, not an evidence source.
# --------------------------------------------------------------------------- #


def _col_letter(col_idx: int) -> str:
    letters = ""
    n = col_idx
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def parse_xlsx_evidence(file_bytes: bytes) -> list[ExtractedChunk]:
    """Extract one `ExtractedChunk` per non-blank row across all sheets, with
    `sheet_name` and a `cell_range` spanning the row's populated columns.
    `page_number` stays null (spreadsheets have no fixed pages).

    Raises `ValueError` if the file cannot be opened as an XLSX workbook, or
    contains no non-blank row on any sheet.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"could not open file as an XLSX workbook: {exc}") from exc

    chunks: list[ExtractedChunk] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=False):
            values = [str(cell.value).strip() for cell in row if cell.value not in (None, "")]
            if not values:
                continue
            first_col = row[0].column
            last_col = row[-1].column
            row_no = row[0].row
            cell_range = (
                f"{_col_letter(first_col)}{row_no}"
                if first_col == last_col
                else f"{_col_letter(first_col)}{row_no}:{_col_letter(last_col)}{row_no}"
            )
            chunks.append(
                ExtractedChunk(
                    sequence_no=len(chunks),
                    text=" | ".join(values),
                    sheet_name=sheet.title,
                    cell_range=cell_range,
                )
            )

    if not chunks:
        raise ValueError("no non-blank rows found in any sheet of the XLSX workbook")

    return chunks


# --------------------------------------------------------------------------- #
# Plain text — one chunk per non-blank line. No page/sheet/heading concept.
# --------------------------------------------------------------------------- #


def parse_plain_text_evidence(file_bytes: bytes) -> list[ExtractedChunk]:
    """Extract one `ExtractedChunk` per non-blank line of decoded text.

    Decoding never raises (`errors="replace"`); raises `ValueError` only if
    every line is blank.
    """
    text = file_bytes.decode("utf-8", errors="replace")
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        raise ValueError("no chunkable text content found in the uploaded document")
    return [ExtractedChunk(sequence_no=i, text=line) for i, line in enumerate(lines)]
