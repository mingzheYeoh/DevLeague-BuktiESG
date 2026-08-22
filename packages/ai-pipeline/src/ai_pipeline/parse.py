"""parse_document() — pure questionnaire parser.

First Vertical Slice scope: a single `.xlsx` questionnaire. No DB session, no
HTTP client, no persistence. Input is raw bytes in, structured data out.

Expected sheet layout (first sheet, header row 1):

    | external_question_id | question_text | section | is_required |

Column names are matched case-insensitively; order in the file does not matter,
only the header text does. Missing optional columns default sensibly.
"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from .models import ParsedQuestion, ParsedQuestionnaire

_REQUIRED_HEADERS = {"external_question_id", "question_text"}

_TRUE_STRINGS = {"true", "yes", "y", "1", "required"}


def _col_letter(col_idx: int) -> str:
    """1-based column index -> spreadsheet column letter (1 -> A, 27 -> AA)."""
    letters = ""
    n = col_idx
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _coerce_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in _TRUE_STRINGS


def parse_document(file_bytes: bytes, filename: str) -> ParsedQuestionnaire:
    """Parse an in-memory `.xlsx` questionnaire into question rows.

    `question_order` is assigned strictly from workbook traversal order — sheet
    order, then row order — never derived from `external_question_id` or any
    other display string (SPEC-AMD-007).

    Raises `ValueError` on a missing required header or an empty question file.
    This function performs no I/O beyond parsing the bytes handed to it: no
    filesystem access, no network call, no database session.
    """

    workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)

    questions: list[ParsedQuestion] = []
    order = 0

    for sheet in workbook.worksheets:
        rows_iter = sheet.iter_rows(values_only=False)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        header_index: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell.value is None:
                continue
            header_index[str(cell.value).strip().lower()] = idx

        missing = _REQUIRED_HEADERS - header_index.keys()
        if missing:
            # Not a questionnaire sheet (or malformed) — skip rather than fail
            # the whole file, unless it's the only/first sheet.
            if sheet is workbook.worksheets[0]:
                raise ValueError(
                    f"{filename}: sheet '{sheet.title}' is missing required "
                    f"header(s): {sorted(missing)}"
                )
            continue

        qid_idx = header_index["external_question_id"]
        text_idx = header_index["question_text"]
        section_idx = header_index.get("section")
        required_idx = header_index.get("is_required")

        for row in rows_iter:
            qid_cell = row[qid_idx] if qid_idx < len(row) else None
            text_cell = row[text_idx] if text_idx < len(row) else None

            qid_value = qid_cell.value if qid_cell is not None else None
            text_value = text_cell.value if text_cell is not None else None

            if qid_value in (None, "") and text_value in (None, ""):
                continue  # skip fully blank row

            section_value = None
            if section_idx is not None and section_idx < len(row):
                section_cell = row[section_idx]
                section_value = (
                    str(section_cell.value) if section_cell.value is not None else None
                )

            required_value = False
            if required_idx is not None and required_idx < len(row):
                required_value = _coerce_bool(row[required_idx].value)

            source_location = f"{sheet.title}!{_col_letter(text_idx + 1)}{text_cell.row}"

            questions.append(
                ParsedQuestion(
                    external_question_id=str(qid_value) if qid_value is not None else "",
                    question_text=str(text_value) if text_value is not None else "",
                    section=section_value,
                    is_required=required_value,
                    source_location=source_location,
                    question_order=order,
                )
            )
            order += 1

    if not questions:
        raise ValueError(f"{filename}: no question rows found")

    return ParsedQuestionnaire(filename=filename, questions=questions)
