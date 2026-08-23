"""Tests for the evidence-document parsers (parse_pdf_evidence,
parse_docx_evidence, parse_xlsx_evidence, parse_plain_text_evidence).

Runs fully offline / deterministic — no live LLM call, no database, no
network (BLOCKER-08).
"""

from __future__ import annotations

import io

import fitz
import pytest
from docx import Document as DocxDocument
from openpyxl import Workbook

from ai_pipeline import (
    parse_docx_evidence,
    parse_pdf_evidence,
    parse_plain_text_evidence,
    parse_xlsx_evidence,
)


def _build_pdf(pages: list[str]) -> bytes:
    doc = fitz.open()
    for text in pages:
        page = doc.new_page()
        if text:
            page.insert_text((72, 72), text)
    buf = doc.tobytes()
    doc.close()
    return buf


def test_parse_pdf_evidence_one_chunk_per_page_with_page_number():
    pdf_bytes = _build_pdf(
        [
            "Total electricity consumption: 12,840 kWh in January 2025.",
            "Waste diverted from landfill: 42 tonnes.",
            "",  # blank page — skipped
        ]
    )

    chunks = parse_pdf_evidence(pdf_bytes)

    assert len(chunks) == 2
    assert chunks[0].page_number == 1
    assert "electricity" in chunks[0].text.lower()
    assert chunks[1].page_number == 2
    assert "waste" in chunks[1].text.lower()
    # sequence_no is 0-based and independent of page_number
    assert [c.sequence_no for c in chunks] == [0, 1]


def test_parse_pdf_evidence_rejects_unparseable_bytes():
    with pytest.raises(ValueError):
        parse_pdf_evidence(b"not a real pdf file")


def test_parse_pdf_evidence_rejects_all_blank_pages():
    pdf_bytes = _build_pdf(["", ""])
    with pytest.raises(ValueError):
        parse_pdf_evidence(pdf_bytes)


def _build_docx(structure: list[tuple[str, str | None]]) -> bytes:
    """`structure` is a list of (text, style_name_or_None) pairs, in order."""
    doc = DocxDocument()
    for text, style in structure:
        doc.add_paragraph(text, style=style)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def test_parse_docx_evidence_chunks_by_heading_section():
    docx_bytes = _build_docx(
        [
            ("Environmental Policy", "Heading 1"),
            ("Total electricity consumption: 12,840 kWh in 2025.", None),
            ("Waste diverted from landfill: 42 tonnes.", None),
            ("Governance", "Heading 1"),
            ("Anti-Bribery", "Heading 2"),
            ("The company maintains a zero-tolerance anti-bribery policy.", None),
        ]
    )

    chunks = parse_docx_evidence(docx_bytes)

    assert len(chunks) == 2
    # DOCX has no fixed pages.
    assert all(c.page_number is None for c in chunks)

    env_chunk = chunks[0]
    assert env_chunk.heading_path == ["Environmental Policy"]
    assert "electricity" in env_chunk.text.lower()
    assert "waste" in env_chunk.text.lower()

    gov_chunk = chunks[1]
    assert gov_chunk.heading_path == ["Governance", "Anti-Bribery"]
    assert "anti-bribery" in gov_chunk.text.lower()


def test_parse_docx_evidence_rejects_unparseable_bytes():
    with pytest.raises(ValueError):
        parse_docx_evidence(b"not a real docx file")


def test_parse_docx_evidence_rejects_empty_document():
    docx_bytes = _build_docx([])
    with pytest.raises(ValueError):
        parse_docx_evidence(docx_bytes)


def test_parse_xlsx_evidence_one_chunk_per_row_with_sheet_and_cell_range():
    wb = Workbook()
    ws = wb.active
    ws.title = "UtilityBill"
    ws.append(["Month", "Electricity (kWh)"])
    ws.append(["January 2025", 12840])
    buf = io.BytesIO()
    wb.save(buf)

    chunks = parse_xlsx_evidence(buf.getvalue())

    assert len(chunks) == 2
    assert chunks[0].sheet_name == "UtilityBill"
    assert chunks[0].cell_range == "A1:B1"
    assert chunks[0].page_number is None
    assert chunks[1].cell_range == "A2:B2"
    assert "12840" in chunks[1].text


def test_parse_xlsx_evidence_handles_rows_narrower_than_the_sheet():
    """A wide table with a narrow note row under it must not crash.

    Regression: the range was previously derived from `row[0]`/`row[-1]`, but
    `read_only=True` pads a short row with `EmptyCell`, which has no `.column`
    or `.row`. Any spreadsheet with a ragged row therefore raised
    AttributeError and surfaced as a 500 on upload. This is the ordinary shape
    of a real waste or utility log, not a corrupt file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "FY2025"
    ws.append(["Month", "Total waste (kg)", "Recycled (kg)", "Rate (%)"])
    ws.append(["2025-01", 18400, 7360, 40])
    ws.append([])  # fully blank row: skipped
    ws.append(["Note", "FY2025 recycling rate 41% per this tracker."])  # narrower
    buf = io.BytesIO()
    wb.save(buf)

    chunks = parse_xlsx_evidence(buf.getvalue())

    assert [c.cell_range for c in chunks] == ["A1:D1", "A2:D2", "A4:B4"]
    assert "41%" in chunks[-1].text
    assert all(c.sheet_name == "FY2025" for c in chunks)


def test_parse_xlsx_evidence_handles_a_row_with_an_empty_leading_column():
    """A row whose first cell is blank must be located from its first
    *populated* cell, not from column A."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Indented"
    ws.append(["Header A", "Header B"])
    ws.append([None, "value in B only"])
    buf = io.BytesIO()
    wb.save(buf)

    chunks = parse_xlsx_evidence(buf.getvalue())

    assert [c.cell_range for c in chunks] == ["A1:B1", "B2"]
    assert chunks[1].text == "value in B only"


def test_parse_xlsx_evidence_rejects_unparseable_bytes():
    with pytest.raises(ValueError):
        parse_xlsx_evidence(b"not a real xlsx file")


def test_parse_plain_text_evidence_one_chunk_per_nonblank_line():
    text = "Line one.\n\nLine two.\n   \nLine three.\n"
    chunks = parse_plain_text_evidence(text.encode("utf-8"))

    assert [c.text for c in chunks] == ["Line one.", "Line two.", "Line three."]
    assert all(c.page_number is None for c in chunks)


def test_parse_plain_text_evidence_rejects_all_blank():
    with pytest.raises(ValueError):
        parse_plain_text_evidence(b"   \n\n   ")
